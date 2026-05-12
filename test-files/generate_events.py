"""
Генератор тестовых событий для ТехНорм.
Формат столбцов соответствует парсеру Excel в DocumentParserService.cs
"""
import random
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

random.seed(42)

# ── Справочники ───────────────────────────────────────────────────────────────

PRODUCTS = [
    ("ВП",  "Вал приводной"),
    ("КР",  "Корпус редуктора"),
    ("ШК",  "Шестерня коническая"),
    ("ФС",  "Фланец соединительный"),
    ("КП",  "Крышка подшипника"),
]

OPERATIONS = [
    ("Заготовительная",         "Участок заготовки",      10, 25),
    ("Токарная (черновая)",     "Токарный станок ТС-16",  20, 50),
    ("Токарная (чистовая)",     "Токарный станок ТС-16",  15, 40),
    ("Фрезерование",            "Фрезерный станок ФС-32", 25, 60),
    ("Сверление",               "Сверлильный станок СС-12", 10, 30),
    ("Нарезание резьбы",        "Сверлильный станок СС-12", 8,  20),
    ("Термическая обработка",   "Термопечь ТП-5",         60, 120),
    ("Шлифование",              "Шлифовальный станок ШС-8", 20, 45),
    ("Контроль размеров",       "ОТК",                    10, 20),
    ("Приёмочный контроль",     "ОТК",                    5,  15),
]

# Для каждого изделия — свой набор технологических операций (индексы в OPERATIONS)
PRODUCT_OPS = {
    "ВП": [0, 1, 2, 7, 8, 9],          # Вал: заготовка → токарка → шлифовка → ОТК
    "КР": [0, 3, 4, 5, 6, 8, 9],       # Корпус: заготовка → фрезер → сверление → термо → ОТК
    "ШК": [0, 1, 2, 6, 7, 8, 9],       # Шестерня: токарка → термо → шлифовка → ОТК
    "ФС": [0, 1, 4, 5, 7, 8, 9],       # Фланец: токарка → сверление → шлифовка → ОТК
    "КП": [0, 1, 2, 4, 8, 9],          # Крышка: токарка → сверление → ОТК
}

START_DATE = date(2025, 1, 1)
END_DATE   = date(2025, 12, 31)

def rand_date():
    delta = END_DATE - START_DATE
    return START_DATE + timedelta(days=random.randint(0, delta.days))

def rand_shift():
    return random.choices([1, 2, 3], weights=[60, 30, 10])[0]


# ── Генерация строк ───────────────────────────────────────────────────────────

rows = []
for code, name in PRODUCTS:
    op_indices = PRODUCT_OPS[code]
    cases_needed = 20           # 20 дел × 5–7 операций ≥ 100 событий
    case_num = 1

    events_for_product = 0
    while events_for_product < 100:
        case_id = f"{code}-{case_num:03d}"
        case_date = rand_date()
        qty = random.randint(1, 5)

        for op_idx in op_indices:
            op_name, resource, dur_min, dur_max = OPERATIONS[op_idx]
            dur = round(random.uniform(dur_min, dur_max), 1)
            # Каждая следующая операция — чуть позже по дате
            op_date = case_date + timedelta(days=op_indices.index(op_idx))
            rows.append({
                "product_name": name,
                "case_id":   case_id,
                "activity":  op_name,
                "date":      op_date.strftime("%d.%m.%Y"),
                "shift":     rand_shift(),
                "resource":  resource,
                "duration":  dur,
                "qty":       qty,
            })
            events_for_product += 1

        case_num += 1

print(f"Всего строк: {len(rows)}")
from collections import Counter
by_product = Counter(r["product_name"] for r in rows)
for p, cnt in sorted(by_product.items()):
    print(f"  {p}: {cnt} событий")


# ── Запись в Excel ────────────────────────────────────────────────────────────

wb = Workbook()
ws = wb.active
ws.title = "События"

HEADER = ["№", "ID дела", "Операция", "Дата", "Смена", "Изделие", "Ресурс", "Длит., мин", "Кол-во"]
COL_WIDTHS = [6, 14, 28, 14, 8, 26, 30, 12, 8]

# Стиль заголовка
hdr_fill   = PatternFill("solid", fgColor="2563EB")
hdr_font   = Font(bold=True, color="FFFFFF", size=10)
hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_side  = Side(style="thin", color="D1D5DB")
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

ws.row_dimensions[1].height = 28
for ci, (h, w) in enumerate(zip(HEADER, COL_WIDTHS), start=1):
    cell = ws.cell(row=1, column=ci, value=h)
    cell.font   = hdr_font
    cell.fill   = hdr_fill
    cell.alignment = hdr_align
    cell.border = thin_border
    ws.column_dimensions[cell.column_letter].width = w

# Цвета строк по изделию
PRODUCT_COLORS = {
    "Вал приводной":          "EFF6FF",
    "Корпус редуктора":       "F0FDF4",
    "Шестерня коническая":    "FFFBEB",
    "Фланец соединительный":  "FDF4FF",
    "Крышка подшипника":      "FFF1F2",
}

data_align  = Alignment(vertical="center")
num_align   = Alignment(horizontal="center", vertical="center")

for ri, row in enumerate(rows, start=2):
    color = PRODUCT_COLORS.get(row["product_name"], "FFFFFF")
    row_fill = PatternFill("solid", fgColor=color)

    values = [
        ri - 1,
        row["case_id"],
        row["activity"],
        row["date"],
        row["shift"],
        row["product_name"],
        row["resource"],
        row["duration"],
        row["qty"],
    ]
    for ci, val in enumerate(values, start=1):
        cell = ws.cell(row=ri, column=ci, value=val)
        cell.fill   = row_fill
        cell.border = thin_border
        cell.alignment = num_align if ci in (1, 4, 5, 8, 9) else data_align

# Заморозить строку заголовка
ws.freeze_panes = "A2"

# Автофильтр
ws.auto_filter.ref = f"A1:I{len(rows)+1}"

out_path = "/home/rnurmanov/myFiles/diplom/project/TechNormBlazor/test-files/events_test.xlsx"
wb.save(out_path)
print(f"\nФайл сохранён: {out_path}")
